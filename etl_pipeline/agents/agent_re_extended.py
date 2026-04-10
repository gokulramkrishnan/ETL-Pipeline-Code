# =============================================================================
# agent_re_extended.py
# Author  : gokulram.krishnan
# Version : 1.0.0
# Purpose : Standalone reverse-engineer agent for DataStage XML, SQL,
#           VQL (Denodo/Vertica), ATL (Attunity/Qlik), and VW (view DDL).
#
# Outputs :
#   BRD.md           Business Requirements Document
#   TRD.md           Technical Requirements Document
#   STTM.xlsx        Source-to-Target Mapping workbook
#   dbt/             Ready-to-run dbt project (staging + raw_vault)
#   lineage.json     Column-level lineage graph
#
# Free LLM enrichment (optional — works without it):
#   Set GROQ_API_KEY  for Groq/llama3-8b  (free tier, fast)
#   Set OLLAMA_URL    for local Ollama     (fully offline)
#   Neither set       → deterministic template output (no API needed)
# =============================================================================

from __future__ import annotations

import json
import os
import re
import sys
import textwrap
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
import xml.etree.ElementTree as ET


# ─────────────────────────────────────────────────────────────────────────────
# 1. FILE PARSERS
# ─────────────────────────────────────────────────────────────────────────────

class ParsedArtifact:
    """Normalised output from any parser."""
    def __init__(self, source_file: str, file_type: str):
        self.source_file   : str        = source_file
        self.file_type     : str        = file_type   # xml|sql|vql|atl|vw
        self.tables        : list[dict] = []   # {name, schema, columns:[{name,type,pk,nullable}]}
        self.views         : list[dict] = []   # {name, schema, sql, columns}
        self.mappings      : list[dict] = []   # {src_table,src_col,tgt_table,tgt_col,expression}
        self.transformations: list[dict]= []   # {name,type,expression}
        self.jobs          : list[str]  = []   # job/procedure names
        self.raw_sql_blocks: list[str]  = []   # raw SQL for LLM context
        self.metadata      : dict       = {}   # extra parser-specific info

    def to_dict(self) -> dict:
        return {k: v for k, v in self.__dict__.items()}


# ── 1a. DataStage XML parser ─────────────────────────────────────────────────

def parse_datastage_xml(path: str) -> ParsedArtifact:
    art = ParsedArtifact(path, "xml")
    try:
        root = ET.parse(path).getroot()
    except ET.ParseError as e:
        art.metadata["error"] = str(e)
        return art

    for job in (root.findall(".//Job") or root.findall(".//DSJob") or [root]):
        job_name = job.get("Identifier") or job.get("name") or "UNKNOWN_JOB"
        art.jobs.append(job_name)

        stages = job.findall(".//Stage") or job.findall(".//DSStage")
        for stage in stages:
            stype = (stage.get("StageType") or stage.get("type") or "").lower()
            sname = stage.get("Identifier") or stage.get("name") or "UNKNOWN"
            cols  = _extract_ds_columns(stage)
            tname = stage.get("TableName") or _prop(stage, "TableName") or sname
            entry = {"name": tname, "schema": "", "columns": cols, "stage_type": stype}
            if any(k in stype for k in ("source","seq","odbc","db2","oracle","jdbc")):
                art.tables.append(entry)
            elif any(k in stype for k in ("target","sink","write")):
                entry["is_target"] = True
                art.tables.append(entry)

        for link in job.findall(".//Link") or []:
            src = link.get("SourceStage","")
            tgt = link.get("TargetStage","")
            for col in _extract_ds_columns(link):
                art.mappings.append({
                    "src_table": src, "src_col": col["name"],
                    "tgt_table": tgt, "tgt_col": col["name"],
                    "expression": col.get("expression","")
                })

    return art


def _extract_ds_columns(elem) -> list[dict]:
    cols = []
    for c in (elem.findall(".//Column") + elem.findall(".//DSColumn")
              + elem.findall(".//column")):
        cols.append({
            "name":       c.get("Identifier") or c.get("name") or "",
            "type":       c.get("SqlType")    or c.get("DataType") or "VARCHAR",
            "nullable":   c.get("Nullable","true").lower() != "false",
            "pk":         False,
            "expression": c.get("Derivation") or c.get("expression") or "",
        })
    return cols


def _prop(elem, name: str) -> str:
    node = elem.find(f".//Property[@name='{name}']")
    return node.text.strip() if node is not None and node.text else ""


# ── 1b. SQL parser ──────────────────────────────────────────────────────────

_SQL_COMMENT    = re.compile(r"--[^\n]*|/\*.*?\*/", re.DOTALL)
_CREATE_TABLE   = re.compile(
    r"CREATE\s+(?:OR\s+REPLACE\s+)?(?:TEMP(?:ORARY)?\s+)?TABLE\s+"
    r"(?:IF\s+NOT\s+EXISTS\s+)?([`\"\[]?[\w\.]+[`\"\]]?)\s*\(([^;]+?)\)",
    re.IGNORECASE | re.DOTALL
)
_CREATE_VIEW    = re.compile(
    r"CREATE\s+(?:OR\s+REPLACE\s+)?VIEW\s+([`\"\[]?[\w\.]+[`\"\]]?)"
    r"\s+AS\s+(SELECT\s+.+?)(?=CREATE|$)",
    re.IGNORECASE | re.DOTALL
)
_INSERT_INTO    = re.compile(
    r"INSERT\s+(?:INTO\s+)?([`\"\[]?[\w\.]+[`\"\]]?)"
    r"(?:\s*\(([^)]+)\))?\s+SELECT\s+(.+?)\s+FROM\s+([`\"\[]?[\w\., ]+[`\"\]]?)"
    r"(?:\s+(?:WHERE|JOIN|LEFT|RIGHT|INNER|OUTER|ON|GROUP|ORDER|LIMIT|;).*)?",
    re.IGNORECASE | re.DOTALL
)
_COL_DEF        = re.compile(
    r"([`\"\[]?[\w]+[`\"\]]?)\s+([\w\(\), ]+?)(?:\s+(?:NOT\s+)?NULL)?"
    r"(?:\s+(?:PRIMARY\s+KEY|UNIQUE|REFERENCES\s+\w+))?(?:\s*,|$)",
    re.IGNORECASE
)


def parse_sql(path: str) -> ParsedArtifact:
    art = ParsedArtifact(path, "sql")
    raw = Path(path).read_text(errors="replace")
    clean = _SQL_COMMENT.sub(" ", raw)
    art.raw_sql_blocks.append(raw[:4000])   # first 4k for LLM context

    for m in _CREATE_TABLE.finditer(clean):
        tname = m.group(1).strip("`\"[]")
        cols  = _parse_col_block(m.group(2))
        art.tables.append({"name": tname, "schema": "", "columns": cols})

    for m in _CREATE_VIEW.finditer(clean):
        vname  = m.group(1).strip("`\"[]")
        vsql   = m.group(2).strip()
        cols   = _infer_select_cols(vsql)
        art.views.append({"name": vname, "schema": "", "sql": vsql, "columns": cols})
        art.raw_sql_blocks.append(vsql[:2000])

    for m in _INSERT_INTO.finditer(clean):
        tgt_table = m.group(1).strip("`\"[]")
        tgt_cols  = [c.strip().strip("`\"[]") for c in m.group(2).split(",")] if m.group(2) else []
        src_list  = m.group(3).strip()
        src_table = m.group(4).strip().split(",")[0].strip().strip("`\"[]")
        for i, sc in enumerate(src_list.split(",")[:len(tgt_cols)]):
            sc = sc.strip().strip("`\"[]")
            art.mappings.append({
                "src_table": src_table, "src_col": sc,
                "tgt_table": tgt_table, "tgt_col": tgt_cols[i] if i < len(tgt_cols) else sc,
                "expression": sc,
            })

    return art


def _parse_col_block(block: str) -> list[dict]:
    cols = []
    for line in block.split("\n"):
        line = line.strip().rstrip(",")
        if not line or line.upper().startswith(("PRIMARY","UNIQUE","KEY","INDEX","CONSTRAINT","CHECK")):
            continue
        m = _COL_DEF.match(line)
        if m:
            cname = m.group(1).strip("`\"[]")
            ctype = m.group(2).strip()
            cols.append({
                "name":     cname,
                "type":     ctype,
                "nullable": "NOT NULL" not in line.upper(),
                "pk":       "PRIMARY KEY" in line.upper(),
                "expression": "",
            })
    return cols


def _infer_select_cols(sql: str) -> list[dict]:
    """Best-effort column extraction from SELECT clause."""
    sel = re.search(r"SELECT\s+(.+?)\s+FROM", sql, re.IGNORECASE | re.DOTALL)
    if not sel:
        return []
    cols = []
    for token in sel.group(1).split(","):
        token = token.strip()
        alias = re.search(r"(?:AS\s+)?([`\"\[]?[\w]+[`\"\]]?)$", token, re.IGNORECASE)
        name  = alias.group(1).strip("`\"[]") if alias else token.split(".")[-1].strip("`\"[]")
        cols.append({"name": name, "type": "VARCHAR", "nullable": True, "pk": False, "expression": token})
    return cols


# ── 1c. VQL parser (Denodo/Vertica Query Language) ──────────────────────────

_VQL_CREATE_VIEW  = re.compile(
    r"CREATE\s+(?:OR\s+REPLACE\s+)?(?:INTERFACE\s+)?VIEW\s+([\w\.]+)"
    r"\s+AS\s+SELECT\s+(.+?)\s+FROM\s+([\w\.]+)(.*?)(?=CREATE|$)",
    re.IGNORECASE | re.DOTALL
)
_VQL_DATASOURCE   = re.compile(
    r"CREATE\s+(?:OR\s+REPLACE\s+)?(?:JDBC|DF|WS|CUSTOM|LDAP|ODBC|WEB|BSON|JSON|XML|ITP|SAPERP)?\s*"
    r"DATA\s*SOURCE\s+([\w]+)\s+(.*?)(?=CREATE|$)",
    re.IGNORECASE | re.DOTALL
)
_VQL_BASE_VIEW    = re.compile(
    r"CREATE\s+(?:OR\s+REPLACE\s+)?BASE\s+VIEW\s+([\w\.]+)\s*\((.+?)\)",
    re.IGNORECASE | re.DOTALL
)


def parse_vql(path: str) -> ParsedArtifact:
    art = ParsedArtifact(path, "vql")
    raw = Path(path).read_text(errors="replace")
    clean = _SQL_COMMENT.sub(" ", raw)
    art.raw_sql_blocks.append(raw[:4000])

    for m in _VQL_BASE_VIEW.finditer(clean):
        vname = m.group(1)
        cols  = _parse_col_block(m.group(2))
        art.tables.append({"name": vname, "schema": "denodo", "columns": cols})

    for m in _VQL_CREATE_VIEW.finditer(clean):
        vname   = m.group(1)
        sel_sql = m.group(2)
        src     = m.group(3)
        vsql    = f"SELECT {sel_sql} FROM {src}{m.group(4)}"
        cols    = _infer_select_cols(vsql)
        art.views.append({"name": vname, "schema": "denodo", "sql": vsql, "columns": cols})
        for col in cols:
            art.mappings.append({
                "src_table":  src,
                "src_col":    col["expression"].split(".")[-1].strip("`\"[] "),
                "tgt_table":  vname,
                "tgt_col":    col["name"],
                "expression": col["expression"],
            })

    for m in _VQL_DATASOURCE.finditer(clean):
        art.metadata.setdefault("data_sources", []).append(m.group(1))

    return art


# ── 1d. ATL parser (Attunity / Qlik Replicate Task Language) ────────────────
# ATL is JSON-like or XML — we handle both

_ATL_TABLE_RULE = re.compile(
    r'"table_name"\s*:\s*"([^"]+)".*?"schema_name"\s*:\s*"([^"]*)"',
    re.DOTALL
)
_ATL_COL_RULE   = re.compile(
    r'"column_name"\s*:\s*"([^"]+)".*?"data_type"\s*:\s*"([^"]*)"',
    re.DOTALL
)
_ATL_TRANSFORM  = re.compile(
    r'"expression"\s*:\s*"([^"]+)".*?"target_column"\s*:\s*"([^"]*)"',
    re.DOTALL
)


def parse_atl(path: str) -> ParsedArtifact:
    art = ParsedArtifact(path, "atl")
    raw = Path(path).read_text(errors="replace")
    art.raw_sql_blocks.append(raw[:4000])

    # Try JSON first (Qlik ATL is JSON)
    try:
        data = json.loads(raw)
        _parse_atl_json(data, art)
        return art
    except json.JSONDecodeError:
        pass

    # Try XML (Attunity ATL is XML)
    try:
        root = ET.fromstring(raw)
        _parse_atl_xml(root, art)
        return art
    except ET.ParseError:
        pass

    # Fallback: regex heuristic
    for m in _ATL_TABLE_RULE.finditer(raw):
        art.tables.append({"name": m.group(1), "schema": m.group(2), "columns": []})
    for m in _ATL_TRANSFORM.finditer(raw):
        art.transformations.append({"name": m.group(2), "type": "expression", "expression": m.group(1)})

    return art


def _parse_atl_json(data: dict | list, art: ParsedArtifact) -> None:
    """Recursively walk Qlik ATL JSON structure."""
    if isinstance(data, list):
        for item in data:
            _parse_atl_json(item, art)
        return
    if not isinstance(data, dict):
        return

    # Task-level table mapping
    for key in ("tableRules", "table_rules", "tables"):
        if key in data:
            for tbl in data[key]:
                name   = tbl.get("tableName", tbl.get("table_name", "UNKNOWN"))
                schema = tbl.get("schemaName", tbl.get("schema_name", ""))
                cols   = []
                for col in tbl.get("columns", tbl.get("columnRules", [])):
                    cols.append({
                        "name":       col.get("columnName", col.get("column_name", "")),
                        "type":       col.get("dataType",   col.get("data_type", "VARCHAR")),
                        "nullable":   col.get("nullable", True),
                        "pk":         col.get("isPrimaryKey", False),
                        "expression": col.get("expression", ""),
                    })
                art.tables.append({"name": name, "schema": schema, "columns": cols})

    for k, v in data.items():
        if isinstance(v, (dict, list)):
            _parse_atl_json(v, art)


def _parse_atl_xml(root, art: ParsedArtifact) -> None:
    """Parse Attunity/CDC ATL XML structure."""
    for tbl in root.findall(".//*[@TableName]") + root.findall(".//*[@table_name]"):
        name   = tbl.get("TableName") or tbl.get("table_name", "UNKNOWN")
        schema = tbl.get("Schema") or tbl.get("schema", "")
        cols   = []
        for col in tbl.findall(".//*[@ColumnName]") + tbl.findall(".//*[@column_name]"):
            cols.append({
                "name":       col.get("ColumnName") or col.get("column_name", ""),
                "type":       col.get("DataType")   or col.get("data_type", "VARCHAR"),
                "nullable":   (col.get("Nullable","Y") or "Y").upper() != "N",
                "pk":         (col.get("IsPK","N")     or "N").upper() == "Y",
                "expression": col.get("Expression") or "",
            })
        art.tables.append({"name": name, "schema": schema, "columns": cols})
    for expr in root.findall(".//*[@Expression]"):
        art.transformations.append({
            "name":       expr.get("TargetColumn", "UNKNOWN"),
            "type":       "expression",
            "expression": expr.get("Expression",""),
        })


# ── 1e. VW (CREATE VIEW DDL) parser ─────────────────────────────────────────

def parse_vw(path: str) -> ParsedArtifact:
    """
    .vw files are typically a single CREATE VIEW statement.
    Delegates to parse_sql after forcing the view pattern.
    """
    art = parse_sql(path)
    art.file_type = "vw"
    # If nothing found as a view, treat entire file body as a view definition
    if not art.views and not art.tables:
        raw = Path(path).read_text(errors="replace")
        stem = Path(path).stem
        cols = _infer_select_cols(raw)
        art.views.append({"name": stem, "schema": "", "sql": raw, "columns": cols})
    return art


# ── 1f. Router ───────────────────────────────────────────────────────────────

def parse_file(path: str) -> ParsedArtifact:
    """Dispatch to the correct parser based on file extension."""
    ext = Path(path).suffix.lower()
    dispatch = {
        ".xml": parse_datastage_xml,
        ".sql": parse_sql,
        ".vql": parse_vql,
        ".atl": parse_atl,
        ".vw":  parse_vw,
    }
    parser = dispatch.get(ext, parse_sql)
    return parser(path)


def merge_artifacts(artifacts: list[ParsedArtifact]) -> dict:
    """Merge all parsed artifacts into a single unified metadata dict."""
    all_tables = []
    all_views  = []
    all_maps   = []
    all_trans  = []
    seen_tables: set[str] = set()

    for art in artifacts:
        for t in art.tables:
            key = (t["name"].upper(), art.source_file)
            if key not in seen_tables:
                t["_source_file"] = art.source_file
                t["_file_type"]   = art.file_type
                all_tables.append(t)
                seen_tables.add(key)
        for v in art.views:
            v["_source_file"] = art.source_file
            all_views.append(v)
        for m in art.mappings:
            m["_source_file"] = art.source_file
            all_maps.append(m)
        all_trans.extend(art.transformations)

    return {
        "tables":          all_tables,
        "views":           all_views,
        "mappings":        all_maps,
        "transformations": all_trans,
        "source_files":    [a.source_file for a in artifacts],
        "parsed_at":       datetime.now(timezone.utc).isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 2. FREE LLM ENRICHMENT  (optional — pipeline works without it)
# ─────────────────────────────────────────────────────────────────────────────

def _call_groq(prompt: str, model: str = "llama3-8b-8192") -> str:
    """Call Groq free-tier API. Set GROQ_API_KEY env var."""
    import urllib.request
    key = os.environ.get("GROQ_API_KEY", "")
    if not key:
        return ""
    body = json.dumps({
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "max_tokens": 1200,
        "temperature": 0.2,
    }).encode()
    req = urllib.request.Request(
        "https://api.groq.com/openai/v1/chat/completions",
        data=body,
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())["choices"][0]["message"]["content"]
    except Exception:
        return ""


def _call_ollama(prompt: str, model: str = "llama3") -> str:
    """Call local Ollama. Set OLLAMA_URL (default http://localhost:11434)."""
    import urllib.request
    base = os.environ.get("OLLAMA_URL", "http://localhost:11434")
    body = json.dumps({"model": model, "prompt": prompt, "stream": False}).encode()
    req  = urllib.request.Request(
        f"{base}/api/generate",
        data=body,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read()).get("response", "")
    except Exception:
        return ""


def llm_enrich(prompt: str) -> str:
    """Try Groq → Ollama → empty string (graceful degradation)."""
    result = _call_groq(prompt)
    if result:
        return result.strip()
    result = _call_ollama(prompt)
    if result:
        return result.strip()
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# 3. OUTPUT GENERATORS
# ─────────────────────────────────────────────────────────────────────────────

# ── 3a. BRD ─────────────────────────────────────────────────────────────────

def generate_brd(meta: dict, out_dir: Path, use_llm: bool = True) -> Path:
    """Business Requirements Document in Markdown."""
    tables  = meta["tables"]
    views   = meta["views"]
    sources = meta["source_files"]
    ts      = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    llm_summary = ""
    if use_llm and (tables or views):
        ctx = f"Tables: {[t['name'] for t in tables[:8]]}. Views: {[v['name'] for v in views[:5]]}."
        prompt = (
            f"You are a data architect. Given this ETL metadata, write 2–3 concise business "
            f"requirement paragraphs covering: what business domain this serves, what business "
            f"problems it solves, and data quality expectations. Metadata: {ctx}"
        )
        llm_summary = llm_enrich(prompt)

    brd_summary = llm_summary or (
        f"This ETL pipeline processes data across {len(tables)} source/target table(s) "
        f"and {len(views)} view(s) sourced from {len(sources)} input file(s). "
        f"The pipeline supports analytical and reporting workloads by transforming "
        f"operational data into a structured, query-optimised form."
    )

    tbl_rows = "\n".join(
        f"| `{t['name']}` | `{t.get('schema') or 'default'}` | "
        f"{len(t.get('columns',[]))} | {t.get('_file_type','').upper()} |"
        for t in tables
    )
    view_rows = "\n".join(
        f"| `{v['name']}` | `{v.get('schema') or 'default'}` | "
        f"{len(v.get('columns',[]))} | `{v.get('sql','')[:80].replace(chr(10),' ')}...` |"
        for v in views
    )

    md = f"""# Business Requirements Document (BRD)

| Field | Value |
|-------|-------|
| Author | gokulram.krishnan |
| Generated | {ts} |
| Source files | {', '.join(Path(s).name for s in sources)} |
| Version | 1.0 |

---

## 1. Business Context

{brd_summary}

---

## 2. Scope

### 2.1 In Scope

- Extraction and parsing of {len(sources)} source artifact(s)
- Transformation logic captured from {len(meta.get('mappings',[]))} column-level mapping(s)
- {len(tables)} table(s) and {len(views)} view(s) identified

### 2.2 Out of Scope

- Physical deployment to production Snowflake
- Scheduler / orchestration tooling configuration
- Data masking beyond column-level PII flagging

---

## 3. Source Inventory

| Table / Object | Schema | Columns | Source Type |
|----------------|--------|---------|-------------|
{tbl_rows or "_No tables detected_"}

### Views

| View Name | Schema | Columns | SQL Preview |
|-----------|--------|---------|-------------|
{view_rows or "_No views detected_"}

---

## 4. Business Rules

| # | Rule | Source |
|---|------|--------|
{"".join(f"| {i+1} | {t.get('expression','DIRECT LOAD')} | {t.get('name','')} |{chr(10)}" for i, t in enumerate(meta.get('transformations',[])[:20])) or "| 1 | No explicit transformation rules detected — direct column mapping assumed | All sources |\n"}

---

## 5. Data Quality Expectations

| Dimension | Requirement |
|-----------|-------------|
| Completeness | All primary key columns must be non-null |
| Uniqueness | No duplicate primary key values in target |
| Timeliness | Pipeline run within defined SLA window |
| Accuracy | Numeric aggregates reconciled source-to-target within 0.001% |
| Consistency | Reference values validated against lookup tables |

---

## 6. Stakeholder Sign-Off

| Role | Name | Date |
|------|------|------|
| Business Owner | | |
| Data Architect | gokulram.krishnan | {ts} |
| Data Engineer | | |
"""
    out = out_dir / "BRD.md"
    out.write_text(md)
    return out


# ── 3b. TRD ─────────────────────────────────────────────────────────────────

def generate_trd(meta: dict, out_dir: Path, use_llm: bool = True) -> Path:
    """Technical Requirements Document in Markdown."""
    tables  = meta["tables"]
    views   = meta["views"]
    sources = meta["source_files"]
    maps    = meta["mappings"]
    ts      = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    llm_tech = ""
    if use_llm and tables:
        ctx     = f"Tables: {[t['name'] for t in tables[:6]]}. Mappings: {len(maps)}."
        prompt  = (
            f"You are a senior data engineer. Given this ETL metadata ({ctx}), write a concise "
            f"technical design paragraph covering: recommended loading strategy (full/incremental), "
            f"hash key approach (MD5/SHA-256), Data Vault 2.0 entity classification rationale, "
            f"and Snowflake-specific optimisations (clustering keys, transient tables)."
        )
        llm_tech = llm_enrich(prompt)

    tech_design = llm_tech or (
        "The pipeline uses a full-load strategy for the initial load, transitioning to "
        "incremental CDC-based loading post-stabilisation. Hash keys are generated using MD5 "
        "on the business key columns. Entities are classified into Data Vault 2.0 Hubs (business "
        "keys), Links (relationships), and Satellites (descriptive attributes). Snowflake "
        "incremental materialisation with `unique_key` is used for Raw Vault tables."
    )

    # Column-level mapping table
    map_rows = "\n".join(
        f"| `{m['src_table']}` | `{m['src_col']}` | `{m['tgt_table']}` "
        f"| `{m['tgt_col']}` | `{m.get('expression','DIRECT')}` |"
        for m in maps[:40]
    )

    # Table schema sections
    schema_sections = ""
    for t in tables[:15]:
        col_rows = "\n".join(
            f"| `{c['name']}` | `{c['type']}` "
            f"| {'✓' if c.get('pk') else ''} "
            f"| {'NULL' if c.get('nullable', True) else 'NOT NULL'} "
            f"| `{c.get('expression','') or 'DIRECT'}` |"
            for c in t.get("columns", [])
        )
        if col_rows:
            schema_sections += f"""
### `{t['name']}`  _(source: {Path(t.get('_source_file','')).name})_

| Column | Data Type | PK | Nullable | Expression |
|--------|-----------|----|----------|------------|
{col_rows}

"""

    md = f"""# Technical Requirements Document (TRD)

| Field | Value |
|-------|-------|
| Author | gokulram.krishnan |
| Generated | {ts} |
| Pipeline version | 1.0 |
| Target platform | Snowflake + dbt + datavault4dbt v1.17.0 |

---

## 1. Technical Architecture

### 1.1 Processing Design

{tech_design}

### 1.2 Layer Architecture

| Layer | Snowflake Schema | dbt Materialisation | Description |
|-------|-----------------|---------------------|-------------|
| Staging | `STAGING` | `view` | Raw column selection, hash key computation |
| Raw Vault | `RAW_VAULT` | `incremental` | Hubs, Links, Satellites |
| Business Vault | `BUSINESS_VAULT` | `table` | Aggregated, business-friendly views |

### 1.3 Hash Key Strategy

- **Algorithm:** MD5 (configurable to SHA-256)
- **Hub HK pattern:** `MD5(CAST(<business_key> AS VARCHAR))`
- **Hash diff pattern:** `MD5(CONCAT_WS('||', <all_descriptor_columns>))`

---

## 2. Source File Inventory

| File | Type | Tables | Views | Mappings |
|------|------|--------|-------|----------|
{"".join(f'| `{Path(s).name}` | `{Path(s).suffix.upper()}` | - | - | - |{chr(10)}' for s in sources)}

---

## 3. Table Schema Definitions

{schema_sections or "_No column-level schema detected._"}

---

## 4. Column-Level Mapping (first 40)

| Source Table | Source Column | Target Table | Target Column | Expression |
|--------------|---------------|--------------|---------------|------------|
{map_rows or "_No explicit column mappings detected._"}

---

## 5. Technical Constraints

| Constraint | Detail |
|-----------|--------|
| Null handling | `NVL(<col>, 0)` for numeric; `COALESCE(<col>, '')` for varchar |
| Date format | ISO 8601 — `YYYY-MM-DD HH:MM:SS.fff` |
| Encoding | UTF-8 throughout |
| Max row size | Snowflake default (16 MB compressed) |
| Batch window | Off-peak 02:00–06:00 UTC |

---

## 6. Error Handling

| Scenario | Action |
|----------|--------|
| Null in PK column | Reject row, write to `REJECTED_RECORDS` table |
| Duplicate PK | Log warning, keep latest by `LOAD_DATE` |
| Schema drift | Alert via dbt test failure; halt pipeline |
| Source unavailable | Retry 3× with exponential back-off, then alert |

---

## 7. Performance Targets

| Metric | Target |
|--------|--------|
| Full load volume | Up to 100M rows |
| Incremental load time | ≤ 15 min per run |
| dbt model compile time | ≤ 60 sec |
| Query SLA (BI layer) | p95 ≤ 5 sec |
"""
    out = out_dir / "TRD.md"
    out.write_text(md)
    return out


# ── 3c. STTM (Excel) ─────────────────────────────────────────────────────────

def generate_sttm(meta: dict, out_dir: Path) -> Path:
    """Source-to-Target Mapping workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback: write CSV
        out = out_dir / "STTM.csv"
        rows = ["Source Table,Source Column,Target Table,Target Column,Expression,PII,Notes"]
        for m in meta["mappings"]:
            rows.append(f'{m["src_table"]},{m["src_col"]},{m["tgt_table"]},{m["tgt_col"]},{m.get("expression","DIRECT")},,')
        out.write_text("\n".join(rows))
        return out

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Mappings ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Mappings"
    hdr_fill = PatternFill("solid", fgColor="0D1B3E")
    hdr_font = Font(bold=True, color="FFFFFF")
    headers  = ["#", "Source File", "Source Table", "Source Column",
                "Target Table", "Target Column", "Expression",
                "Transformation Type", "PII Flag", "Notes"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = hdr_fill; cell.font = hdr_font

    _PII_NAMES = {"email","phone","ssn","dob","passport","credit_card","ip_address","name","address"}
    for i, m in enumerate(meta["mappings"], 1):
        expr  = m.get("expression","") or "DIRECT"
        ttype = ("DIRECT" if expr in ("DIRECT","",m["src_col"])
                 else "HASH" if "MD5" in expr.upper() or "SHA" in expr.upper()
                 else "EXPRESSION")
        pii   = "PII" if any(p in m["tgt_col"].lower() for p in _PII_NAMES) else ""
        ws1.append([i, Path(m.get("_source_file","")).name,
                    m["src_table"], m["src_col"],
                    m["tgt_table"], m["tgt_col"],
                    expr, ttype, pii, ""])
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 3

    # ── Sheet 2: Table Inventory ───────────────────────────────────────────
    ws2 = wb.create_sheet("Table Inventory")
    ws2.append(["Table Name", "Schema", "Source File", "Type", "Column Count", "PK Columns"])
    for cell in ws2[1]: cell.fill = hdr_fill; cell.font = hdr_font
    for t in meta["tables"]:
        pks = ", ".join(c["name"] for c in t.get("columns",[]) if c.get("pk"))
        ws2.append([t["name"], t.get("schema",""), Path(t.get("_source_file","")).name,
                    t.get("_file_type","").upper(), len(t.get("columns",[])), pks])

    # ── Sheet 3: Column Schema ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Column Schema")
    ws3.append(["Table", "Column", "Data Type", "PK", "Nullable", "Expression", "PII"])
    for cell in ws3[1]: cell.fill = hdr_fill; cell.font = hdr_font
    for t in meta["tables"]:
        for c in t.get("columns",[]):
            pii = "PII" if any(p in c["name"].lower() for p in _PII_NAMES) else ""
            ws3.append([t["name"], c["name"], c.get("type","VARCHAR"),
                        "YES" if c.get("pk") else "",
                        "NULL" if c.get("nullable",True) else "NOT NULL",
                        c.get("expression",""), pii])

    # ── Sheet 4: Views ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Views")
    ws4.append(["View Name", "Schema", "Source File", "Column Count", "SQL (first 500 chars)"])
    for cell in ws4[1]: cell.fill = hdr_fill; cell.font = hdr_font
    for v in meta["views"]:
        ws4.append([v["name"], v.get("schema",""), Path(v.get("_source_file","")).name,
                    len(v.get("columns",[])), (v.get("sql","") or "")[:500]])

    # ── Sheet 5: Summary ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("Summary")
    ws5.append(["Metric", "Value"])
    ws5.append(["Total Tables",      len(meta["tables"])])
    ws5.append(["Total Views",        len(meta["views"])])
    ws5.append(["Total Mappings",     len(meta["mappings"])])
    ws5.append(["Total Transformations", len(meta["transformations"])])
    ws5.append(["Source Files",       len(meta["source_files"])])
    ws5.append(["Generated At",       meta["parsed_at"]])

    out = out_dir / "STTM.xlsx"
    wb.save(out)
    return out


# ── 3d. dbt pipeline ─────────────────────────────────────────────────────────

def generate_dbt(meta: dict, out_dir: Path,
                 project_name: str = "re_pipeline",
                 schema_stg: str = "STAGING",
                 schema_raw: str = "RAW_VAULT") -> Path:
    """Generate a ready-to-run minimal dbt project."""
    dbt = out_dir / "dbt"
    models_stg = dbt / "models" / "staging"
    models_raw = dbt / "models" / "raw_vault"
    for d in [models_stg, models_raw]:
        d.mkdir(parents=True, exist_ok=True)

    # packages.yml
    (dbt / "packages.yml").write_text(
        "packages:\n"
        "  - package: ScalefreeCOM/datavault4dbt\n"
        "    version: [\">=1.17.0\", \"<2.0.0\"]\n"
    )

    # dbt_project.yml
    (dbt / "dbt_project.yml").write_text(textwrap.dedent(f"""\
        name: '{project_name}'
        version: '1.0.0'
        config-version: 2
        profile: '{project_name}'

        vars:
          datavault4dbt:
            hash:          'MD5'
            load_datetime: 'LOAD_DATE'
            record_source: 'RECORD_SOURCE'

        models:
          {project_name}:
            staging:
              +schema: {schema_stg}
              +materialized: view
            raw_vault:
              +schema: {schema_raw}
              +materialized: incremental
              +unique_key: HK
    """))

    # profiles.yml
    (dbt / "profiles.yml").write_text(textwrap.dedent(f"""\
        {project_name}:
          target: dev
          outputs:
            dev:
              type: snowflake
              account: "{{{{ env_var('SNOWFLAKE_ACCOUNT') }}}}"
              user:    "{{{{ env_var('SNOWFLAKE_USER') }}}}"
              password:"{{{{ env_var('SNOWFLAKE_PASSWORD') }}}}"
              role:    "{{{{ env_var('SNOWFLAKE_ROLE', 'TRANSFORMER') }}}}"
              database:"{{{{ env_var('SNOWFLAKE_DATABASE') }}}}"
              warehouse:"{{{{ env_var('SNOWFLAKE_WAREHOUSE') }}}}"
              schema:  {schema_stg}
              threads: 4
    """))

    # sources.yml
    _src_tables = meta["tables"] + [
        {"name": v["name"], "schema": v.get("schema",""), "columns": v.get("columns",[])}
        for v in meta["views"]
    ]
    src_block = "version: 2\nsources:\n  - name: raw_source\n    tables:\n"
    for t in _src_tables:
        src_block += f"      - name: {t['name']}\n"
    (models_stg / "sources.yml").write_text(src_block)

    # Staging models — one per source table
    for t in meta["tables"]:
        cols_block = "\n        ".join(
            f"MD5(CAST({c['name']} AS VARCHAR)) AS {c['name'].upper().rstrip('_ID')}_HK,"
            if c.get("pk") else f"{c['name']},"
            for c in t.get("columns", [])
        ).rstrip(",")
        name = re.sub(r"[^a-zA-Z0-9_]", "_", t["name"].lower())
        sql = textwrap.dedent(f"""\
            {{{{ config(materialized='view') }}}}

            SELECT
                {cols_block or "    *"},
                CURRENT_TIMESTAMP()        AS LOAD_DATE,
                '{Path(t.get('_source_file','')).name}' AS RECORD_SOURCE
            FROM {{{{ source('raw_source', '{t['name']}') }}}}
        """)
        (models_stg / f"stg_{name}.sql").write_text(sql)

    # Hub models — one per table with a PK
    for t in meta["tables"]:
        pk_cols = [c for c in t.get("columns",[]) if c.get("pk")]
        if not pk_cols:
            continue
        entity = re.sub(r"^(SRC|STG|TGT|DIM|FACT|RAW)_", "", t["name"].upper())
        hk_col = f"{entity}_HK"
        bk_col = pk_cols[0]["name"]
        stg_name = re.sub(r"[^a-zA-Z0-9_]", "_", t["name"].lower())
        sql = textwrap.dedent(f"""\
            {{{{ config(materialized='incremental', unique_key='{hk_col}') }}}}

            {{{{ datavault4dbt.hub(
                src_pk='{hk_col}',
                src_bk=['{bk_col}'],
                src_ldts='LOAD_DATE',
                src_source='RECORD_SOURCE',
                source_model='stg_{stg_name}'
            ) }}}}
        """)
        (models_raw / f"hub_{entity.lower()}.sql").write_text(sql)

    # Satellite models — descriptive columns
    for t in meta["tables"]:
        pk_cols   = [c for c in t.get("columns",[]) if c.get("pk")]
        desc_cols = [c for c in t.get("columns",[]) if not c.get("pk")]
        if not pk_cols or not desc_cols:
            continue
        entity   = re.sub(r"^(SRC|STG|TGT|DIM|FACT|RAW)_", "", t["name"].upper())
        hk_col   = f"{entity}_HK"
        payload  = "\n        ".join(f"- {c['name']}" for c in desc_cols[:20])
        stg_name = re.sub(r"[^a-zA-Z0-9_]", "_", t["name"].lower())
        sql = textwrap.dedent(f"""\
            {{{{ config(materialized='incremental', unique_key='{hk_col}') }}}}

            {{{{ datavault4dbt.sat(
                src_pk='{hk_col}',
                src_hashdiff={{is_hashdiff: true, columns: [{', '.join(repr(c['name']) for c in desc_cols[:10])}]}},
                src_payload=[{', '.join(repr(c['name']) for c in desc_cols[:20])}],
                src_ldts='LOAD_DATE',
                src_source='RECORD_SOURCE',
                source_model='stg_{stg_name}'
            ) }}}}
        """)
        (models_raw / f"sat_{entity.lower()}.sql").write_text(sql)

    # schema.yml for tests
    test_yaml = "version: 2\nmodels:\n"
    for t in meta["tables"]:
        pk_cols = [c for c in t.get("columns",[]) if c.get("pk")]
        entity  = re.sub(r"^(SRC|STG|TGT|DIM|FACT|RAW)_", "", t["name"].upper())
        if pk_cols:
            test_yaml += f"  - name: hub_{entity.lower()}\n    columns:\n"
            test_yaml += f"      - name: {entity}_HK\n        tests: [not_null, unique]\n"
    (models_raw / "schema.yml").write_text(test_yaml)

    return dbt


# ── 3e. Lineage ───────────────────────────────────────────────────────────────

def generate_lineage(meta: dict, out_dir: Path) -> Path:
    """Column-level lineage as JSON + Mermaid flowchart."""
    nodes = {}
    edges = []

    for t in meta["tables"]:
        nid = t["name"]
        nodes[nid] = {"id": nid, "type": "table", "schema": t.get("schema",""), "columns": [c["name"] for c in t.get("columns",[])]}
    for v in meta["views"]:
        nid = v["name"]
        nodes[nid] = {"id": nid, "type": "view", "schema": v.get("schema",""), "columns": [c["name"] for c in v.get("columns",[])]}
    for m in meta["mappings"]:
        edges.append({
            "from_table":  m["src_table"],
            "from_column": m["src_col"],
            "to_table":    m["tgt_table"],
            "to_column":   m["tgt_col"],
            "expression":  m.get("expression",""),
        })

    lineage = {"nodes": list(nodes.values()), "edges": edges, "generated_at": meta["parsed_at"]}

    json_out = out_dir / "lineage.json"
    with open(json_out, "w") as f:
        json.dump(lineage, f, indent=2, default=str)

    # Mermaid flowchart (table-level, max 20 edges to keep diagram readable)
    mmd_lines = ["flowchart LR"]
    seen_nodes: set[str] = set()
    for e in edges[:30]:
        src = re.sub(r"[^a-zA-Z0-9]","_", e["from_table"])
        tgt = re.sub(r"[^a-zA-Z0-9]","_", e["to_table"])
        if src not in seen_nodes:
            mmd_lines.append(f'    {src}["{e["from_table"]}"]')
            seen_nodes.add(src)
        if tgt not in seen_nodes:
            mmd_lines.append(f'    {tgt}["{e["to_table"]}"]')
            seen_nodes.add(tgt)
        label = e["from_column"] if len(e["from_column"]) < 25 else e["from_column"][:22]+"..."
        mmd_lines.append(f'    {src} -->|"{label}"| {tgt}')

    mmd_out = out_dir / "lineage.mmd"
    mmd_out.write_text("\n".join(mmd_lines))

    return json_out


# ─────────────────────────────────────────────────────────────────────────────
# 4. MAIN AGENT CLASS
# ─────────────────────────────────────────────────────────────────────────────

class REExtendedAgent:
    """
    Standalone Reverse Engineer — Extended

    Parses DataStage XML, SQL, VQL, ATL, VW files and produces:
      BRD.md        Business Requirements Document
      TRD.md        Technical Requirements Document
      STTM.xlsx     Source-to-Target Mapping workbook
      dbt/          Ready-to-run dbt project
      lineage.json  Column-level lineage graph
      lineage.mmd   Mermaid flowchart

    Usage:
        agent = REExtendedAgent()
        result = agent.run(
            input_files = ["job.xml", "transforms.sql", "views.vw"],
            output_dir  = "output/re_extended",
            use_llm     = True,   # set False for fully offline operation
        )
    """

    def run(
        self,
        input_files: list[str],
        output_dir:  str  = "output/re_extended",
        use_llm:     bool = True,
        project_name:str  = "re_pipeline",
    ) -> dict:
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)

        print(f"  [RE-Extended] Parsing {len(input_files)} file(s)...")
        artifacts = []
        for f in input_files:
            p = Path(f)
            if not p.exists():
                print(f"  [RE-Extended] ⚠️  Not found: {f}")
                continue
            art = parse_file(str(p))
            artifacts.append(art)
            print(f"  [RE-Extended]   ✅ {p.name}: {len(art.tables)} tables, "
                  f"{len(art.views)} views, {len(art.mappings)} mappings")

        if not artifacts:
            print("  [RE-Extended] ❌ No valid files parsed")
            return {"error": "no valid input files"}

        meta  = merge_artifacts(artifacts)
        llm_ok = bool(os.environ.get("GROQ_API_KEY") or os.environ.get("OLLAMA_URL"))
        if use_llm and not llm_ok:
            print("  [RE-Extended] ℹ️  No LLM key set — using deterministic templates (set GROQ_API_KEY or OLLAMA_URL for enrichment)")

        print("  [RE-Extended] Generating BRD.md...")
        brd = generate_brd(meta, out, use_llm=use_llm and llm_ok)

        print("  [RE-Extended] Generating TRD.md...")
        trd = generate_trd(meta, out, use_llm=use_llm and llm_ok)

        print("  [RE-Extended] Generating STTM.xlsx...")
        sttm = generate_sttm(meta, out)

        print("  [RE-Extended] Generating dbt pipeline...")
        dbt_dir = generate_dbt(meta, out, project_name=project_name)

        print("  [RE-Extended] Generating lineage.json...")
        lineage = generate_lineage(meta, out)

        # Save unified metadata for the confidence comparator
        meta_path = out / "unified_metadata.json"
        with open(meta_path, "w") as f:
            json.dump(meta, f, indent=2, default=str)

        print(f"  [RE-Extended] ✅ All outputs written to {out}")
        return {
            "output_dir": str(out),
            "files": {
                "brd":      str(brd),
                "trd":      str(trd),
                "sttm":     str(sttm),
                "dbt":      str(dbt_dir),
                "lineage":  str(lineage),
                "metadata": str(meta_path),
            },
            "summary": {
                "tables":      len(meta["tables"]),
                "views":       len(meta["views"]),
                "mappings":    len(meta["mappings"]),
                "dbt_staging": len(list((dbt_dir/"models"/"staging").glob("stg_*.sql"))),
                "dbt_hubs":    len(list((dbt_dir/"models"/"raw_vault").glob("hub_*.sql"))),
                "dbt_sats":    len(list((dbt_dir/"models"/"raw_vault").glob("sat_*.sql"))),
            },
        }


# ─────────────────────────────────────────────────────────────────────────────
# 5. CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def _cli():
    import argparse
    parser = argparse.ArgumentParser(
        prog="agent_re_extended",
        description="Reverse-engineer DataStage XML, SQL, VQL, ATL, VW → BRD, TRD, STTM, dbt, lineage",
    )
    parser.add_argument("files",    nargs="+", help="Input files (.xml .sql .vql .atl .vw)")
    parser.add_argument("--out",    default="output/re_extended", help="Output directory")
    parser.add_argument("--name",   default="re_pipeline",        help="dbt project name")
    parser.add_argument("--no-llm", action="store_true",          help="Disable LLM enrichment")
    args = parser.parse_args()

    result = REExtendedAgent().run(
        input_files  = args.files,
        output_dir   = args.out,
        use_llm      = not args.no_llm,
        project_name = args.name,
    )
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    _cli()
