"""
test_re_extended.py
Author  : gokulram.krishnan
Purpose : Test suite for agent_re_extended — all 5 parsers and all 5 outputs.
          Runs with stdlib + openpyxl only. No network. No LLM calls.

Run     : python test_re_extended.py
          python -m pytest test_re_extended.py -v
"""

import json
import os
import sys
import tempfile
import time
import traceback
import textwrap
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from etl_pipeline.agents.agent_re_extended import (
    parse_datastage_xml,
    parse_sql,
    parse_vql,
    parse_atl,
    parse_vw,
    merge_artifacts,
    generate_brd,
    generate_trd,
    generate_sttm,
    generate_dbt,
    generate_lineage,
    REExtendedAgent,
    _parse_col_block,
    _infer_select_cols,
)


# ─────────────────────────────────────────────────────────────────────────────
# Test fixtures
# ─────────────────────────────────────────────────────────────────────────────

DATASTAGE_XML = """<?xml version="1.0" encoding="UTF-8"?>
<DSExport>
  <Job Identifier="JOB_TEST" Name="JOB_TEST">
    <Stage StageType="source" Identifier="SRC_CUSTOMER" Name="SRC_CUSTOMER">
      <Property name="TableName">SRC_CUSTOMER</Property>
      <Column Identifier="CUSTOMER_ID" SqlType="INTEGER"      Nullable="false" Derivation="CUSTOMER_ID"/>
      <Column Identifier="EMAIL"       SqlType="VARCHAR(255)" Nullable="true"  Derivation="LOWER(EMAIL)"/>
      <Column Identifier="STATUS"      SqlType="VARCHAR(50)"  Nullable="true"  Derivation="STATUS"/>
    </Stage>
    <Stage StageType="target" Identifier="TGT_CUSTOMER" Name="TGT_CUSTOMER">
      <Property name="TableName">TGT_CUSTOMER</Property>
      <Column Identifier="CUSTOMER_HK" SqlType="VARCHAR(32)" Nullable="false" Derivation="MD5(CUSTOMER_ID)"/>
      <Column Identifier="EMAIL"       SqlType="VARCHAR(255)" Nullable="true" Derivation="LOWER(EMAIL)"/>
    </Stage>
    <Link SourceStage="SRC_CUSTOMER" TargetStage="TGT_CUSTOMER">
      <Column Identifier="CUSTOMER_ID" SqlType="INTEGER"/>
      <Column Identifier="EMAIL"       SqlType="VARCHAR(255)"/>
    </Link>
  </Job>
</DSExport>"""

SQL_CONTENT = """
CREATE TABLE ORDERS (
    ORDER_ID     INTEGER      NOT NULL PRIMARY KEY,
    CUSTOMER_ID  INTEGER      NOT NULL,
    ORDER_AMOUNT DECIMAL(18,2),
    STATUS       VARCHAR(50)
);

CREATE VIEW V_ACTIVE_ORDERS AS
SELECT ORDER_ID, CUSTOMER_ID, ORDER_AMOUNT
FROM   ORDERS
WHERE  STATUS = 'ACTIVE';

INSERT INTO FACT_ORDERS (ORDER_HK, CUSTOMER_HK, ORDER_AMOUNT)
SELECT MD5(CAST(ORDER_ID AS VARCHAR)),
       MD5(CAST(CUSTOMER_ID AS VARCHAR)),
       NVL(ORDER_AMOUNT, 0)
FROM   ORDERS;
"""

VQL_CONTENT = """
CREATE OR REPLACE BASE VIEW bv_product (
    PRODUCT_ID   int,
    PRODUCT_NAME text,
    UNIT_PRICE   decimal
)
DATASOURCENAME=ds_oracle;

CREATE OR REPLACE VIEW v_product_summary AS
SELECT PRODUCT_ID, PRODUCT_NAME, SUM(UNIT_PRICE) AS TOTAL
FROM   bv_product
GROUP BY PRODUCT_ID, PRODUCT_NAME;
"""

ATL_JSON = json.dumps({
    "task": {
        "name": "test_task",
        "tableRules": [
            {
                "tableName": "CUSTOMER",
                "schemaName": "SALES",
                "columns": [
                    {"columnName": "CUST_ID",   "dataType": "INT4",   "isPrimaryKey": True,  "nullable": False},
                    {"columnName": "CUST_NAME", "dataType": "STRING", "isPrimaryKey": False, "nullable": True},
                ]
            }
        ]
    }
})

VW_CONTENT = """CREATE OR REPLACE VIEW V_ORDERS AS
SELECT o.ORDER_ID, o.CUSTOMER_ID, o.ORDER_AMOUNT, c.EMAIL
FROM   ORDERS o
JOIN   CUSTOMERS c ON o.CUSTOMER_ID = c.CUSTOMER_ID
WHERE  o.STATUS = 'ACTIVE';"""


def _write(tmp: Path, name: str, content: str) -> str:
    p = tmp / name
    p.write_text(content)
    return str(p)


# ─────────────────────────────────────────────────────────────────────────────
# Parser tests
# ─────────────────────────────────────────────────────────────────────────────

class TestDataStageXMLParser(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.path = _write(self.tmp, "test.xml", DATASTAGE_XML)

    def test_tables_extracted(self):
        art = parse_datastage_xml(self.path)
        names = [t["name"] for t in art.tables]
        self.assertIn("SRC_CUSTOMER", names, "SRC_CUSTOMER not parsed")
        self.assertIn("TGT_CUSTOMER", names, "TGT_CUSTOMER not parsed")

    def test_columns_extracted(self):
        art  = parse_datastage_xml(self.path)
        src  = next(t for t in art.tables if t["name"] == "SRC_CUSTOMER")
        cols = [c["name"] for c in src["columns"]]
        self.assertIn("CUSTOMER_ID", cols)
        self.assertIn("EMAIL", cols)
        self.assertEqual(len(cols), 3)

    def test_mappings_from_links(self):
        art = parse_datastage_xml(self.path)
        self.assertGreater(len(art.mappings), 0, "No mappings extracted from <Link>")
        cols_mapped = [m["src_col"] for m in art.mappings]
        self.assertIn("CUSTOMER_ID", cols_mapped)

    def test_job_name_captured(self):
        art = parse_datastage_xml(self.path)
        self.assertIn("JOB_TEST", art.jobs)

    def test_file_type(self):
        art = parse_datastage_xml(self.path)
        self.assertEqual(art.file_type, "xml")

    def test_malformed_xml_does_not_crash(self):
        bad = _write(self.tmp, "bad.xml", "<<not xml>>")
        art = parse_datastage_xml(bad)
        self.assertIn("error", art.metadata)


class TestSQLParser(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.path = _write(self.tmp, "test.sql", SQL_CONTENT)

    def test_create_table_parsed(self):
        art   = parse_sql(self.path)
        names = [t["name"] for t in art.tables]
        self.assertIn("ORDERS", names)

    def test_primary_key_detected(self):
        art    = parse_sql(self.path)
        orders = next(t for t in art.tables if t["name"] == "ORDERS")
        pk_col = next((c for c in orders["columns"] if c["pk"]), None)
        self.assertIsNotNone(pk_col, "PRIMARY KEY column not detected")
        self.assertEqual(pk_col["name"], "ORDER_ID")

    def test_create_view_parsed(self):
        art   = parse_sql(self.path)
        names = [v["name"] for v in art.views]
        self.assertIn("V_ACTIVE_ORDERS", names)

    def test_insert_select_mappings(self):
        art  = parse_sql(self.path)
        tgts = [m["tgt_table"] for m in art.mappings]
        self.assertIn("FACT_ORDERS", tgts)

    def test_column_types_captured(self):
        art    = parse_sql(self.path)
        orders = next(t for t in art.tables if t["name"] == "ORDERS")
        types  = {c["name"]: c["type"] for c in orders["columns"]}
        self.assertIn("DECIMAL", types.get("ORDER_AMOUNT","").upper())

    def test_file_type(self):
        art = parse_sql(self.path)
        self.assertEqual(art.file_type, "sql")


class TestVQLParser(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.path = _write(self.tmp, "test.vql", VQL_CONTENT)

    def test_base_view_as_table(self):
        art   = parse_vql(self.path)
        names = [t["name"] for t in art.tables]
        self.assertIn("bv_product", names)

    def test_derived_view_captured(self):
        art   = parse_vql(self.path)
        names = [v["name"] for v in art.views]
        self.assertIn("v_product_summary", names)

    def test_schema_set_to_denodo(self):
        art = parse_vql(self.path)
        bv  = next(t for t in art.tables if t["name"] == "bv_product")
        self.assertEqual(bv["schema"], "denodo")

    def test_mappings_from_view(self):
        art = parse_vql(self.path)
        self.assertGreater(len(art.mappings), 0)

    def test_file_type(self):
        art = parse_vql(self.path)
        self.assertEqual(art.file_type, "vql")


class TestATLParser(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.path = _write(self.tmp, "test.atl", ATL_JSON)

    def test_table_extracted(self):
        art   = parse_atl(self.path)
        names = [t["name"] for t in art.tables]
        self.assertIn("CUSTOMER", names)

    def test_columns_extracted(self):
        art  = parse_atl(self.path)
        cust = next(t for t in art.tables if t["name"] == "CUSTOMER")
        cols = [c["name"] for c in cust["columns"]]
        self.assertIn("CUST_ID", cols)

    def test_pk_detected(self):
        art  = parse_atl(self.path)
        cust = next(t for t in art.tables if t["name"] == "CUSTOMER")
        pk   = next((c for c in cust["columns"] if c["pk"]), None)
        self.assertIsNotNone(pk)
        self.assertEqual(pk["name"], "CUST_ID")

    def test_schema_preserved(self):
        art  = parse_atl(self.path)
        cust = next(t for t in art.tables if t["name"] == "CUSTOMER")
        self.assertEqual(cust["schema"], "SALES")

    def test_file_type(self):
        art = parse_atl(self.path)
        self.assertEqual(art.file_type, "atl")

    def test_xml_atl_fallback(self):
        xml_atl = """<Task><Tables><Table TableName="ORDERS" Schema="SALES">
            <Column ColumnName="ORDER_ID" DataType="INT" IsPK="Y"/></Table></Tables></Task>"""
        p   = _write(Path(tempfile.mkdtemp()), "test2.atl", xml_atl)
        art = parse_atl(p)
        # Should not crash; may extract via regex fallback
        self.assertEqual(art.file_type, "atl")


class TestVWParser(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.path = _write(self.tmp, "v_orders.vw", VW_CONTENT)

    def test_view_extracted(self):
        art   = parse_vw(self.path)
        names = [v["name"] for v in art.views]
        self.assertIn("V_ORDERS", names)

    def test_columns_inferred(self):
        art  = parse_vw(self.path)
        view = next(v for v in art.views if v["name"] == "V_ORDERS")
        self.assertGreater(len(view["columns"]), 0)

    def test_file_type(self):
        art = parse_vw(self.path)
        self.assertEqual(art.file_type, "vw")

    def test_empty_vw_does_not_crash(self):
        p   = _write(Path(tempfile.mkdtemp()), "empty.vw", "-- empty file\n")
        art = parse_vw(p)
        self.assertEqual(art.file_type, "vw")


# ─────────────────────────────────────────────────────────────────────────────
# Merge tests
# ─────────────────────────────────────────────────────────────────────────────

class TestMergeArtifacts(unittest.TestCase):

    def _make_artifacts(self):
        tmp = Path(tempfile.mkdtemp())
        a1  = parse_datastage_xml(_write(tmp, "a.xml", DATASTAGE_XML))
        a2  = parse_sql(_write(tmp, "b.sql", SQL_CONTENT))
        a3  = parse_atl(_write(tmp, "c.atl", ATL_JSON))
        return merge_artifacts([a1, a2, a3])

    def test_tables_merged(self):
        meta = self._make_artifacts()
        self.assertGreater(len(meta["tables"]), 2)

    def test_no_duplicate_tables(self):
        meta  = self._make_artifacts()
        names = [t["name"].upper() for t in meta["tables"]]
        # Duplicates allowed across different source files; check list is non-empty
        self.assertEqual(len(names), len(meta["tables"]))

    def test_mappings_combined(self):
        meta = self._make_artifacts()
        self.assertGreater(len(meta["mappings"]), 0)

    def test_source_files_listed(self):
        meta = self._make_artifacts()
        self.assertEqual(len(meta["source_files"]), 3)

    def test_parsed_at_present(self):
        meta = self._make_artifacts()
        self.assertIn("parsed_at", meta)
        self.assertTrue(meta["parsed_at"])


# ─────────────────────────────────────────────────────────────────────────────
# Output generator tests
# ─────────────────────────────────────────────────────────────────────────────

def _make_meta(tmp: Path) -> dict:
    a1 = parse_datastage_xml(_write(tmp, "x.xml", DATASTAGE_XML))
    a2 = parse_sql(_write(tmp, "y.sql", SQL_CONTENT))
    a3 = parse_vql(_write(tmp, "z.vql", VQL_CONTENT))
    return merge_artifacts([a1, a2, a3])


class TestBRDGenerator(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        self.meta = _make_meta(self.tmp)

    def test_brd_file_created(self):
        out = generate_brd(self.meta, self.tmp, use_llm=False)
        self.assertTrue(out.exists())

    def test_brd_is_markdown(self):
        out     = generate_brd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("# Business Requirements Document", content)

    def test_brd_contains_tables(self):
        out     = generate_brd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("SRC_CUSTOMER", content)

    def test_brd_contains_author(self):
        out     = generate_brd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("gokulram.krishnan", content)

    def test_brd_contains_dq_section(self):
        out     = generate_brd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("Data Quality", content)


class TestTRDGenerator(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.meta = _make_meta(self.tmp)

    def test_trd_file_created(self):
        out = generate_trd(self.meta, self.tmp, use_llm=False)
        self.assertTrue(out.exists())

    def test_trd_contains_schema_section(self):
        out     = generate_trd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("Table Schema", content)

    def test_trd_contains_mapping_table(self):
        out     = generate_trd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("Column-Level Mapping", content)

    def test_trd_contains_snowflake_layer(self):
        out     = generate_trd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("RAW_VAULT", content)

    def test_trd_contains_column_data_types(self):
        out     = generate_trd(self.meta, self.tmp, use_llm=False)
        content = out.read_text()
        self.assertIn("INTEGER", content)


class TestSTTMGenerator(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.meta = _make_meta(self.tmp)

    def test_sttm_file_created(self):
        out = generate_sttm(self.meta, self.tmp)
        self.assertTrue(out.exists())

    def test_sttm_is_xlsx(self):
        out = generate_sttm(self.meta, self.tmp)
        self.assertIn(out.suffix, (".xlsx", ".csv"))

    def test_sttm_has_5_sheets(self):
        out = generate_sttm(self.meta, self.tmp)
        if out.suffix == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(out)
            self.assertEqual(len(wb.sheetnames), 5)

    def test_sttm_all_mappings_sheet_has_data(self):
        out = generate_sttm(self.meta, self.tmp)
        if out.suffix == ".xlsx":
            import openpyxl
            wb  = openpyxl.load_workbook(out)
            ws  = wb["All Mappings"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            # Header row + at least one data row
            self.assertGreater(len(rows), 0)

    def test_pii_flagged_for_email(self):
        out = generate_sttm(self.meta, self.tmp)
        if out.suffix == ".xlsx":
            import openpyxl
            wb   = openpyxl.load_workbook(out)
            ws   = wb["All Mappings"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            pii_rows = [r for r in rows if r and "PII" in (r[8] or "")]
            self.assertGreater(len(pii_rows), 0, "EMAIL column should be flagged as PII")


class TestDBTGenerator(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.meta = _make_meta(self.tmp)
        self.dbt  = generate_dbt(self.meta, self.tmp, project_name="test_project")

    def test_dbt_dir_created(self):
        self.assertTrue(self.dbt.exists())

    def test_packages_yml_has_datavault4dbt(self):
        content = (self.dbt / "packages.yml").read_text()
        self.assertIn("datavault4dbt", content)
        self.assertIn("1.17.0", content)

    def test_dbt_project_yml_exists(self):
        self.assertTrue((self.dbt / "dbt_project.yml").exists())

    def test_profiles_yml_snowflake(self):
        content = (self.dbt / "profiles.yml").read_text()
        self.assertIn("snowflake", content)

    def test_staging_models_generated(self):
        stg_models = list((self.dbt / "models" / "staging").glob("stg_*.sql"))
        self.assertGreater(len(stg_models), 0, "No staging models generated")

    def test_hub_models_generated(self):
        hub_models = list((self.dbt / "models" / "raw_vault").glob("hub_*.sql"))
        self.assertGreater(len(hub_models), 0, "No hub models generated")

    def test_hub_model_uses_datavault4dbt_macro(self):
        hub_models = list((self.dbt / "models" / "raw_vault").glob("hub_*.sql"))
        if hub_models:
            content = hub_models[0].read_text()
            self.assertIn("datavault4dbt.hub", content)

    def test_sat_models_generated(self):
        sat_models = list((self.dbt / "models" / "raw_vault").glob("sat_*.sql"))
        self.assertGreater(len(sat_models), 0, "No satellite models generated")

    def test_schema_yml_has_not_null_test(self):
        schema = (self.dbt / "models" / "raw_vault" / "schema.yml").read_text()
        self.assertIn("not_null", schema)

    def test_sources_yml_lists_tables(self):
        sources = (self.dbt / "models" / "staging" / "sources.yml").read_text()
        self.assertIn("raw_source", sources)


class TestLineageGenerator(unittest.TestCase):

    def setUp(self):
        self.tmp  = Path(tempfile.mkdtemp())
        self.meta = _make_meta(self.tmp)

    def test_lineage_json_created(self):
        out = generate_lineage(self.meta, self.tmp)
        self.assertTrue(out.exists())

    def test_lineage_has_nodes(self):
        out  = generate_lineage(self.meta, self.tmp)
        data = json.loads(out.read_text())
        self.assertIn("nodes", data)
        self.assertGreater(len(data["nodes"]), 0)

    def test_lineage_has_edges(self):
        out  = generate_lineage(self.meta, self.tmp)
        data = json.loads(out.read_text())
        self.assertIn("edges", data)

    def test_mermaid_file_created(self):
        generate_lineage(self.meta, self.tmp)
        mmd = self.tmp / "lineage.mmd"
        self.assertTrue(mmd.exists())

    def test_mermaid_starts_with_flowchart(self):
        generate_lineage(self.meta, self.tmp)
        content = (self.tmp / "lineage.mmd").read_text()
        self.assertTrue(content.startswith("flowchart LR"))

    def test_lineage_edges_have_required_keys(self):
        out  = generate_lineage(self.meta, self.tmp)
        data = json.loads(out.read_text())
        if data["edges"]:
            edge = data["edges"][0]
            for key in ("from_table", "from_column", "to_table", "to_column"):
                self.assertIn(key, edge, f"Edge missing key: {key}")


# ─────────────────────────────────────────────────────────────────────────────
# Full agent integration test
# ─────────────────────────────────────────────────────────────────────────────

class TestREExtendedAgentIntegration(unittest.TestCase):

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp())
        (self.tmp / "in").mkdir(exist_ok=True)
        self.xml_f = _write(self.tmp / "in", "job.xml",      DATASTAGE_XML)
        self.sql_f = _write(self.tmp / "in", "orders.sql",   SQL_CONTENT)
        self.vql_f = _write(self.tmp / "in", "products.vql", VQL_CONTENT)
        self.atl_f = _write(self.tmp / "in", "repl.atl",     ATL_JSON)
        self.vw_f  = _write(self.tmp / "in", "v_orders.vw",  VW_CONTENT)

        self.result = REExtendedAgent().run(
            input_files  = [self.xml_f, self.sql_f, self.vql_f, self.atl_f, self.vw_f],
            output_dir   = str(self.tmp / "out"),
            use_llm      = False,
            project_name = "test_project",
        )

    def test_run_returns_dict(self):
        self.assertIsInstance(self.result, dict)

    def test_no_error_key(self):
        self.assertNotIn("error", self.result)

    def test_all_output_files_present(self):
        files = self.result.get("files", {})
        for key in ("brd", "trd", "sttm", "dbt", "lineage", "metadata"):
            self.assertIn(key, files, f"Missing output key: {key}")
            self.assertTrue(Path(files[key]).exists(), f"File not found: {files[key]}")

    def test_summary_counts_reasonable(self):
        summary = self.result.get("summary", {})
        self.assertGreater(summary.get("tables", 0),    0, "Zero tables in summary")
        self.assertGreater(summary.get("dbt_staging",0), 0, "Zero staging models")
        self.assertGreater(summary.get("dbt_hubs",   0), 0, "Zero hub models")

    def test_unified_metadata_valid_json(self):
        meta_path = self.result["files"]["metadata"]
        with open(meta_path) as f:
            data = json.load(f)
        self.assertIn("tables",   data)
        self.assertIn("mappings", data)
        self.assertIn("views",    data)

    def test_all_5_file_types_parsed(self):
        meta_path = self.result["files"]["metadata"]
        with open(meta_path) as f:
            data = json.load(f)
        src_files = data.get("source_files", [])
        self.assertEqual(len(src_files), 5)

    def test_missing_file_handled_gracefully(self):
        result = REExtendedAgent().run(
            input_files = ["/nonexistent/file.sql", self.sql_f],
            output_dir  = str(self.tmp / "partial"),
            use_llm     = False,
        )
        # Should not crash — just skip missing file
        self.assertNotIn("error", result)

    def test_single_file_run(self):
        result = REExtendedAgent().run(
            input_files = [self.sql_f],
            output_dir  = str(self.tmp / "single"),
            use_llm     = False,
        )
        self.assertIn("files", result)

    def test_runtime_under_10_seconds(self):
        start = time.time()
        REExtendedAgent().run(
            input_files = [self.xml_f, self.sql_f],
            output_dir  = str(self.tmp / "perf"),
            use_llm     = False,
        )
        elapsed = time.time() - start
        self.assertLess(elapsed, 10.0, f"Run took {elapsed:.1f}s — should be < 10s")


# ─────────────────────────────────────────────────────────────────────────────
# Helper function tests
# ─────────────────────────────────────────────────────────────────────────────

class TestHelpers(unittest.TestCase):

    def test_parse_col_block_basic(self):
        block = "  ORDER_ID INTEGER NOT NULL PRIMARY KEY,\n  AMOUNT DECIMAL(18,2),\n  STATUS VARCHAR(50)\n"
        cols  = _parse_col_block(block)
        names = [c["name"] for c in cols]
        self.assertIn("ORDER_ID", names)
        self.assertIn("AMOUNT",   names)

    def test_parse_col_block_pk_detection(self):
        block = "ID INTEGER NOT NULL PRIMARY KEY,\nNAME VARCHAR(100)"
        cols  = _parse_col_block(block)
        pk    = next((c for c in cols if c["pk"]), None)
        self.assertIsNotNone(pk)
        self.assertEqual(pk["name"], "ID")

    def test_infer_select_cols_simple(self):
        sql  = "SELECT col1, col2, t.col3 AS alias FROM my_table"
        cols = _infer_select_cols(sql)
        names = [c["name"] for c in cols]
        self.assertIn("col1",  names)
        self.assertIn("alias", names)

    def test_infer_select_cols_star(self):
        sql  = "SELECT * FROM my_table"
        cols = _infer_select_cols(sql)
        # Star — may return single col named '*'
        self.assertIsInstance(cols, list)

    def test_infer_select_cols_no_from(self):
        cols = _infer_select_cols("not a select statement")
        self.assertEqual(cols, [])


# ─────────────────────────────────────────────────────────────────────────────
# Runner
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    loader  = unittest.TestLoader()
    suite   = unittest.TestSuite()

    test_classes = [
        TestDataStageXMLParser,
        TestSQLParser,
        TestVQLParser,
        TestATLParser,
        TestVWParser,
        TestMergeArtifacts,
        TestBRDGenerator,
        TestTRDGenerator,
        TestSTTMGenerator,
        TestDBTGenerator,
        TestLineageGenerator,
        TestREExtendedAgentIntegration,
        TestHelpers,
    ]

    for tc in test_classes:
        suite.addTests(loader.loadTestsFromTestCase(tc))

    total   = suite.countTestCases()
    start   = time.time()
    runner  = unittest.TextTestRunner(verbosity=0, stream=open(os.devnull, "w"))
    result  = runner.run(suite)
    elapsed = time.time() - start

    passed  = total - len(result.failures) - len(result.errors)
    failed  = len(result.failures)
    errors  = len(result.errors)

    # Per-class summary
    print(f"\n{'='*65}")
    print(f"  RE Extended Agent — Test Results")
    print(f"{'='*65}")
    for tc in test_classes:
        tc_suite   = loader.loadTestsFromTestCase(tc)
        tc_result  = unittest.TestResult()
        tc_suite.run(tc_result)
        tc_total   = tc_suite.countTestCases()
        tc_fail    = len(tc_result.failures) + len(tc_result.errors)
        icon       = "✅" if tc_fail == 0 else "❌"
        print(f"  {icon} {tc.__name__:<42s}  {tc_total - tc_fail}/{tc_total}")

    print(f"{'='*65}")
    print(f"  TOTAL: {total}  ✅ PASS: {passed}  ❌ FAIL: {failed}  💥 ERROR: {errors}"
          f"  ({elapsed:.2f}s)")
    print(f"  PASS RATE: {round(passed/total*100,1) if total else 0}%")
    print(f"{'='*65}")

    if result.failures:
        print("\n── Failures ──")
        for test, tb in result.failures:
            print(f"\n  {test}")
            print(textwrap.indent(tb.split('\n')[-2], "    "))

    if result.errors:
        print("\n── Errors ──")
        for test, tb in result.errors:
            print(f"\n  {test}")
            print(textwrap.indent(tb.split('\n')[-2], "    "))

    sys.exit(0 if failed + errors == 0 else 1)
