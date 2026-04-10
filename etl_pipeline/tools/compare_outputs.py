#!/usr/bin/env python3
# =============================================================================
# compare_outputs.py
# Author  : gokulram.krishnan
# Purpose : Compare RE Extended output against other ETL tools / hand-crafted
#           specs to build a confidence matrix.
#
# Comparison dimensions:
#   Table coverage   — tables found vs reference
#   Column coverage  — columns found vs reference
#   Mapping accuracy — expression fidelity
#   View detection   — views found vs reference
#   Lineage depth    — hop count vs reference
#
# Usage:
#   python tools/compare_outputs.py \
#       --our    output/re_extended/unified_metadata.json \
#       --ref    reference/informatica_output.json \
#       --label  "Informatica" \
#       --out    output/confidence_matrix.xlsx
# =============================================================================

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional


# ─────────────────────────────────────────────────────────────────────────────
# Loaders
# ─────────────────────────────────────────────────────────────────────────────

def load_metadata(path: str) -> dict:
    """
    Load metadata JSON.  Accepts:
      - unified_metadata.json from REExtendedAgent
      - unified_metadata.json from ReverseEngineerAgent (existing pipeline)
      - hand-crafted reference JSON (same schema)
    """
    with open(path) as f:
        data = json.load(f)

    # Normalise from ReverseEngineerAgent format if needed
    if "sources" in data and "tables" not in data:
        data["tables"]   = data.pop("sources", [])
        data["views"]    = []
        data["mappings"] = data.get("mappings", [])
        # Normalise column keys
        for t in data["tables"]:
            for c in t.get("columns", []):
                c.setdefault("type",     c.pop("data_type", "VARCHAR"))
                c.setdefault("pk",       c.get("key_type","") == "PRIMARY KEY")
                c.setdefault("nullable", c.get("nullable","Y") != "N")
                c.setdefault("expression", "")
        for m in data.get("mappings", []):
            m.setdefault("src_table",  m.pop("source_table",  ""))
            m.setdefault("src_col",    m.pop("source_column", ""))
            m.setdefault("tgt_table",  m.pop("target_table",  ""))
            m.setdefault("tgt_col",    m.pop("target_column", ""))
    return data


# ─────────────────────────────────────────────────────────────────────────────
# Scoring helpers
# ─────────────────────────────────────────────────────────────────────────────

def _name_set(items: list[dict], key: str = "name") -> set[str]:
    return {str(i.get(key,"")).upper().strip() for i in items if i.get(key)}


def _jaccard(a: set, b: set) -> float:
    if not a and not b:
        return 1.0
    return len(a & b) / len(a | b)


def _precision_recall(our: set, ref: set) -> tuple[float, float]:
    if not our:
        return 0.0, 0.0
    if not ref:
        return 1.0, 1.0
    precision = len(our & ref) / len(our)
    recall    = len(our & ref) / len(ref)
    return precision, recall


def _f1(precision: float, recall: float) -> float:
    if precision + recall == 0:
        return 0.0
    return 2 * precision * recall / (precision + recall)


# ─────────────────────────────────────────────────────────────────────────────
# Dimension scorers
# ─────────────────────────────────────────────────────────────────────────────

def score_table_coverage(our: dict, ref: dict) -> dict:
    our_names = _name_set(our.get("tables", []))
    ref_names = _name_set(ref.get("tables", []))
    p, r      = _precision_recall(our_names, ref_names)
    missing   = sorted(ref_names - our_names)
    extra     = sorted(our_names - ref_names)
    return {
        "dimension":  "Table Coverage",
        "our_count":  len(our_names),
        "ref_count":  len(ref_names),
        "matched":    len(our_names & ref_names),
        "missing":    missing,
        "extra":      extra,
        "precision":  round(p, 3),
        "recall":     round(r, 3),
        "f1":         round(_f1(p,r), 3),
        "jaccard":    round(_jaccard(our_names, ref_names), 3),
    }


def score_column_coverage(our: dict, ref: dict) -> dict:
    """Per-table column coverage, then averaged."""
    our_tables = {t["name"].upper(): t for t in our.get("tables",[])}
    ref_tables = {t["name"].upper(): t for t in ref.get("tables",[])}
    common     = set(our_tables) & set(ref_tables)

    if not common:
        return {"dimension": "Column Coverage", "f1": 0.0, "detail": "no common tables"}

    f1s, details = [], []
    for tname in sorted(common):
        our_cols = _name_set(our_tables[tname].get("columns",[]))
        ref_cols = _name_set(ref_tables[tname].get("columns",[]))
        p, r     = _precision_recall(our_cols, ref_cols)
        f        = _f1(p,r)
        f1s.append(f)
        details.append({"table": tname, "our": len(our_cols), "ref": len(ref_cols),
                         "matched": len(our_cols & ref_cols), "f1": round(f,3)})

    avg_f1 = sum(f1s) / len(f1s) if f1s else 0.0
    return {
        "dimension": "Column Coverage",
        "tables_compared": len(common),
        "avg_f1": round(avg_f1, 3),
        "detail": details,
    }


def score_mapping_accuracy(our: dict, ref: dict) -> dict:
    """
    Mapping accuracy: compare (src_table, src_col, tgt_table, tgt_col) tuples.
    Expressions are fuzzy-matched (normalised whitespace, case-insensitive).
    """
    def _key(m):
        return (m.get("src_table","").upper(), m.get("src_col","").upper(),
                m.get("tgt_table","").upper(), m.get("tgt_col","").upper())

    our_keys = {_key(m) for m in our.get("mappings",[])}
    ref_keys = {_key(m) for m in ref.get("mappings",[])}
    p, r     = _precision_recall(our_keys, ref_keys)

    # Expression fidelity on matched mappings
    our_map  = {_key(m): m for m in our.get("mappings",[])}
    ref_map  = {_key(m): m for m in ref.get("mappings",[])}
    expr_ok  = 0
    for k in our_keys & ref_keys:
        oe = re.sub(r"\s+", " ", (our_map[k].get("expression","") or "DIRECT").upper())
        re_ = re.sub(r"\s+", " ", (ref_map[k].get("expression","") or "DIRECT").upper())
        if oe == re_ or oe == "DIRECT" or re_ == "DIRECT":
            expr_ok += 1

    matched  = len(our_keys & ref_keys)
    expr_pct = round(expr_ok / matched * 100, 1) if matched else 0.0

    return {
        "dimension":         "Mapping Accuracy",
        "our_count":         len(our_keys),
        "ref_count":         len(ref_keys),
        "matched":           matched,
        "precision":         round(p, 3),
        "recall":            round(r, 3),
        "f1":                round(_f1(p,r), 3),
        "expression_match_pct": expr_pct,
    }

import re
import re as _re

def score_view_detection(our: dict, ref: dict) -> dict:
    our_views = _name_set(our.get("views",[]))
    ref_views = _name_set(ref.get("views",[]))
    p, r      = _precision_recall(our_views, ref_views)
    return {
        "dimension": "View Detection",
        "our_count": len(our_views),
        "ref_count": len(ref_views),
        "matched":   len(our_views & ref_views),
        "precision": round(p,3),
        "recall":    round(r,3),
        "f1":        round(_f1(p,r),3),
    }


def score_lineage_depth(our: dict, ref: dict) -> dict:
    """Compare lineage graph hop counts (unique src→tgt table pairs)."""
    def _table_hops(meta):
        return {(m.get("src_table","").upper(), m.get("tgt_table","").upper())
                for m in meta.get("mappings",[])}
    our_hops = _table_hops(our)
    ref_hops = _table_hops(ref)
    p, r     = _precision_recall(our_hops, ref_hops)
    return {
        "dimension":  "Lineage Depth",
        "our_hops":   len(our_hops),
        "ref_hops":   len(ref_hops),
        "matched":    len(our_hops & ref_hops),
        "precision":  round(p,3),
        "recall":     round(r,3),
        "f1":         round(_f1(p,r),3),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Confidence matrix builder
# ─────────────────────────────────────────────────────────────────────────────

DIMENSION_WEIGHTS = {
    "Table Coverage":    0.25,
    "Column Coverage":   0.25,
    "Mapping Accuracy":  0.30,
    "View Detection":    0.10,
    "Lineage Depth":     0.10,
}


def build_confidence_matrix(our_path: str, ref_configs: list[dict]) -> dict:
    """
    ref_configs: [{"path": "...", "label": "Informatica"}, ...]
    """
    our = load_metadata(our_path)
    matrix = []

    for ref_cfg in ref_configs:
        ref   = load_metadata(ref_cfg["path"])
        label = ref_cfg.get("label", Path(ref_cfg["path"]).stem)

        scores = {
            "Table Coverage":   score_table_coverage(our, ref),
            "Column Coverage":  score_column_coverage(our, ref),
            "Mapping Accuracy": score_mapping_accuracy(our, ref),
            "View Detection":   score_view_detection(our, ref),
            "Lineage Depth":    score_lineage_depth(our, ref),
        }

        # Weighted confidence score
        confidence = sum(
            DIMENSION_WEIGHTS[dim] * scores[dim].get("f1", scores[dim].get("avg_f1", 0.0))
            for dim in DIMENSION_WEIGHTS
        )

        matrix.append({
            "tool":       label,
            "confidence": round(confidence * 100, 1),
            "scores":     scores,
        })

    return {
        "our_file":    our_path,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "comparisons": matrix,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Report writers
# ─────────────────────────────────────────────────────────────────────────────

def write_markdown_report(result: dict, out_path: str) -> None:
    lines = [
        "# Confidence Matrix Report",
        "",
        f"**Our output:** `{Path(result['our_file']).name}`  ",
        f"**Generated:** {result['generated_at']}",
        "",
        "## Summary",
        "",
        "| Tool | Confidence % | Table F1 | Column F1 | Mapping F1 | View F1 | Lineage F1 |",
        "|------|-------------|----------|-----------|------------|---------|------------|",
    ]
    for c in result["comparisons"]:
        s = c["scores"]
        lines.append(
            f"| {c['tool']} | **{c['confidence']}%** "
            f"| {s['Table Coverage']['f1']:.2f} "
            f"| {s['Column Coverage'].get('avg_f1', s['Column Coverage'].get('f1',0)):.2f} "
            f"| {s['Mapping Accuracy']['f1']:.2f} "
            f"| {s['View Detection']['f1']:.2f} "
            f"| {s['Lineage Depth']['f1']:.2f} |"
        )
    lines += ["", "## Detail by Tool", ""]
    for c in result["comparisons"]:
        lines.append(f"### {c['tool']} — {c['confidence']}% confidence")
        lines.append("")
        for dim, sc in c["scores"].items():
            lines.append(f"**{dim}**")
            for k, v in sc.items():
                if k == "dimension":
                    continue
                if isinstance(v, list) and len(v) > 5:
                    v = v[:5] + [f"... (+{len(v)-5} more)"]
                lines.append(f"- `{k}`: {v}")
            lines.append("")
    Path(out_path).write_text("\n".join(lines))


def write_excel_report(result: dict, out_path: str) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        write_markdown_report(result, out_path.replace(".xlsx",".md"))
        return

    wb = openpyxl.Workbook()
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    hdr = PatternFill("solid", fgColor="0D1B3E")
    hf  = Font(bold=True, color="FFFFFF")
    headers = ["Tool", "Confidence %", "Table F1", "Column F1",
               "Mapping F1", "View F1", "Lineage F1"]
    ws.append(headers)
    for cell in ws[1]: cell.fill = hdr; cell.font = hf

    for c in result["comparisons"]:
        s = c["scores"]
        cf = c["confidence"]
        ws.append([
            c["tool"], cf,
            s["Table Coverage"]["f1"],
            s["Column Coverage"].get("avg_f1", s["Column Coverage"].get("f1",0)),
            s["Mapping Accuracy"]["f1"],
            s["View Detection"]["f1"],
            s["Lineage Depth"]["f1"],
        ])
        # Colour-code confidence
        row = ws.max_row
        fill = PatternFill("solid", fgColor=(
            "27AE60" if cf >= 80 else "F39C12" if cf >= 50 else "E74C3C"
        ))
        ws.cell(row, 2).fill = fill
        ws.cell(row, 2).font = Font(bold=True, color="FFFFFF")

    # Detail sheet per tool
    for c in result["comparisons"]:
        ws2 = wb.create_sheet(c["tool"][:31])
        ws2.append(["Dimension", "Metric", "Value"])
        for cell in ws2[1]: cell.fill = hdr; cell.font = hf
        for dim, sc in c["scores"].items():
            for k, v in sc.items():
                if k == "dimension": continue
                if isinstance(v, list): v = ", ".join(str(x) for x in v[:10])
                ws2.append([dim, k, str(v)])

    # Add generation metadata
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(["Key","Value"])
    ws_meta.append(["Our file",    result["our_file"]])
    ws_meta.append(["Generated",   result["generated_at"]])
    ws_meta.append(["Dimension weights", str(DIMENSION_WEIGHTS)])

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        prog="compare_outputs",
        description="Compare RE Extended output vs other tools — builds confidence matrix",
    )
    parser.add_argument("--our",    required=True, help="Path to our unified_metadata.json")
    parser.add_argument("--ref",    action="append", nargs=2, metavar=("PATH","LABEL"),
                        help="Reference file and label, e.g. --ref ref.json Informatica (repeatable)")
    parser.add_argument("--out",    default="output/confidence_matrix.xlsx",
                        help="Output path (.xlsx or .md)")
    args = parser.parse_args()

    if not args.ref:
        print("❌ Provide at least one --ref PATH LABEL argument")
        sys.exit(1)

    ref_configs = [{"path": p, "label": lbl} for p, lbl in args.ref]
    result      = build_confidence_matrix(args.our, ref_configs)

    out = args.out
    if out.endswith(".xlsx"):
        write_excel_report(result, out)
        print(f"✅ Excel confidence matrix → {out}")
    else:
        out = out if out.endswith(".md") else out + ".md"
        write_markdown_report(result, out)
        print(f"✅ Markdown confidence matrix → {out}")

    # Also print summary to stdout
    print(f"\n{'─'*50}")
    print(f"  CONFIDENCE MATRIX SUMMARY")
    print(f"{'─'*50}")
    for c in result["comparisons"]:
        bar_len = int(c["confidence"] / 5)
        bar     = "█" * bar_len + "░" * (20 - bar_len)
        print(f"  {c['tool']:20s}  [{bar}]  {c['confidence']:5.1f}%")
    print(f"{'─'*50}")


if __name__ == "__main__":
    main()
